import tkinter as tk
#tkinter._test()
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog

from openpyxl import *
from openpyxl.utils import column_index_from_string
import time
import re

# Guardar como .pyw para no abrir la consola detrás

# -------   FUNCIONES ------------

#global posicion_row_grid
#global posicion_row_grid_destino
#global a_origen

def redimensionar_scrollbar():
	root.update()
	c.config(scrollregion=c.bbox("all"))

def funcion_scroll(event):
	# Para moverse sobre el panel y no solo con las brarras de scroll
	c.scan_mark(event.x, event.y)
	c.scan_dragto(event.x, event.y, gain=1)

def examinar_origen():
	global a_origen

	a_origen = filedialog.askopenfilename(title="Examinar archivo origen", filetypes=(("Excel files","*.xlsx *.xlsm *.xltx *.xltm"),))
	l_t_origen = Label(dobleframe, text="Archivo origen seleccionado:", bg="pale green")
	l_t_origen.grid(row=5, column=0, sticky="e")
	l_a_origen = Label(dobleframe, text= a_origen, bg="pale green")
	l_a_origen.grid(row=6, column=0, columnspan=2, sticky="ew", pady=7)

	redimensionar_scrollbar()

def examinar_destino():
	global posicion_row_grid_destino
	posicion_row_grid_destino = 26
	global a_destino
	global e_sheet_destino 
	global e_fila_destino
	global e_col_ini_destino
	global e_col_fin_destino
	global e_fila_escritura_destino


	a_destino = filedialog.askopenfilename(title="Examinar archivo destino", filetypes=(("Excel files","*.xlsx *.xlsm *.xltx *.xltm"),))
	l_t_destino = Label(dobleframe, text="Archivo destino seleccionado:", bg="pale green")
	l_t_destino.grid(row=17, column=0, sticky="e")
	l_a_destino = Label(dobleframe, text= a_destino, bg="pale green")
	l_a_destino.grid(row=18, column=0, columnspan=2, sticky="ew", pady=6)

	# Indicar cabecero destino:
	l_sheet_destino = Label(dobleframe,text="Indica nombre de pestaña:", bg="pale green")
	l_sheet_destino.grid(row=19, column=0, sticky="e", padx=2, pady=2)
	e_sheet_destino = Entry(dobleframe)
	e_sheet_destino.grid(row=19, column=1, sticky="w", padx=2, pady=2)

	l_fila_destino = Label(dobleframe,text="Indica fila de cabeceras:", bg="pale green")
	l_fila_destino.grid(row=20, column=0, sticky="e", padx=2, pady=2)
	e_fila_destino = Entry(dobleframe)
	e_fila_destino.grid(row=20, column=1, sticky="w", padx=2, pady=2)
	
	l_col_ini_destino = Label(dobleframe,text="Indica columna de inicio:", bg="pale green")
	l_col_ini_destino.grid(row=21, column=0, sticky="e", padx=2, pady=2)
	e_col_ini_destino = Entry(dobleframe)
	e_col_ini_destino.grid(row=21, column=1, sticky="w", padx=2, pady=2)
	
	l_col_fin_destino = Label(dobleframe,text="Indica columna de fin:", bg="pale green")
	l_col_fin_destino.grid(row=22, column=0, sticky="e", padx=2, pady=2)
	e_col_fin_destino = Entry(dobleframe)
	e_col_fin_destino.grid(row=22, column=1, sticky="w", padx=2, pady=2)
	
	l_fila_destino = Label(dobleframe,text="Indica la fila en la que quieres que empiece a escribir los nuevos datos:", bg="pale green")
	l_fila_destino.grid(row=23, column=0, sticky="e", padx=2, pady=2)
	e_fila_escritura_destino = Entry(dobleframe)
	e_fila_escritura_destino.grid(row=23, column=1, sticky="w", padx=2, pady=2)

	b_mostrar_destino = Button(dobleframe, text="Mostrar destino", command=mostrar_destino )
	b_mostrar_destino.grid(row=24, column=1, sticky="e", padx=2, pady=4)	




	redimensionar_scrollbar()

# Para trabajar con los archivos

def revisar_columnas():

	# ORIGEN ----------------

	global posicion_row_grid

	# worbook origen

	if a_origen == None:
		tk.messagebox.showwarning(title="Sin archivo origen", message="Por favor selecciona un archivo de origen")

	elif e_fila_origen == None or e_col_ini_origen == None or e_col_fin_origen == None :
		tk.messagebox.showwarning(title="Sin datos origen", message="Por favor selecciona la fila y las columnas del archivo de origen")

	#elif a_destino != None and (e_fila_destino == None or e_col_ini_destino == None or e_col_fin_destino == None) :
		#tkinter.messagebox.showwarning(title="Sin datos origen", message="Por favor selecciona la fila y las columnas del archivo de destino")

	else :
		
		global a_origen_ok
		a_origen_ok = str(a_origen.replace('\\', '\\\\'))
		#print (a_origen)
		#print (a_origen_ok)
		wb_origen = load_workbook(filename = a_origen_ok)
		# comprobar que exista la pestaña indicada
		pestanas = wb_origen.sheetnames
		if str(e_sheet_origen.get()) in pestanas:
			#sheet_wb_origen = wb_origen.active
			global sheet_wb_origen
			sheet_wb_origen = wb_origen[str(e_sheet_origen.get())]
			fila_origen=str(e_fila_origen.get())
			global col_ini_origen
			col_ini_origen=str(e_col_ini_origen.get())
			global col_fin_origen
			col_fin_origen=str(e_col_fin_origen.get())
			#print(str(fila_origen) + " - "+ str(col_ini_origen) + " - "+ str(col_fin_origen))
			#global celda_ini_origen
			celda_ini_origen = str(col_ini_origen + fila_origen)
			#global celda_fin_origen
			celda_fin_origen = str(col_fin_origen + fila_origen)
			global rango_celdas_origen
			rango_celdas_origen = str(celda_ini_origen + ":" + celda_fin_origen)
			#print(rango_celdas_origen)
			tupla_origen = sheet_wb_origen[rango_celdas_origen]
			global dic_origen
			dic_origen = {}

			global lista_cab_origen 
			lista_cab_origen = []
	
			# listar cabecera de origen
	
			posicion_row_grid = 28
	
			#print (tupla_origen)
			# lista_col_origen=[] # No necesario
			for elementos_tupla_origen in tupla_origen :
				#print (elementos_tupla_origen)
				#lista_col_origen.append(elementos_tupla_origen) # No necesario
				#print (lista_col_origen) # No necesario
				# salida: [(<Cell 'Hoja1'.A1>, <Cell 'Hoja1'.B1>, <Cell 'Hoja1'.C1>)]
				for val_col_origen in elementos_tupla_origen:
					# celda origen
					#print (val_col_origen)
					# salida: <Cell 'Hoja1'.A1>
					pos_origen = re.search(r"\.(\w+)>$", str(val_col_origen))
					pos_cel_origen = pos_origen.group(1)
					lista_cab_origen.append(pos_cel_origen) 
					#pos_col_origen = val_col_origen.column # Las enumera en vez de poner sus letras
					#pos_fil_origen = val_col_origen.row
					#pos_cel_origen = str(pos_col_origen + pos_fil_origen)
					l_pos_col_origen = Label(dobleframe,text=pos_cel_origen, bg="pale green", wraplength=30)
					l_pos_col_origen.grid(row=posicion_row_grid, column=0, sticky="ew", padx=2, pady=2)
					# valores origen
					#print (val_col_origen.value)
					# salida: Col-1
					l_val_col_origen = Label(dobleframe,text=val_col_origen.value, bg="pale green", wraplength=70)
					l_val_col_origen.grid(row=posicion_row_grid, column=1, sticky="w", padx=2, pady=2)

					dic_origen.update({pos_cel_origen:val_col_origen.value})
					#print(dic_origen)

					posicion_row_grid = posicion_row_grid + 1

		else:
			tk.messagebox.showwarning(title="Origen", message="Por favor revisa el nombre de la pestaña de origen.")

		redimensionar_scrollbar()

		# DESTINO  ----------

		# Nuevo Destino 

		if v_cb_nuevo_destino.get() is True and v_cb_destino.get() is False:

			#print ("Detecta el check como true")

			b_nuevo_destino = Button(dobleframe, text="Confirmar", command=nuevo_destino )
			b_nuevo_destino.grid(row=posicion_row_grid + 1, column=0, sticky="e", padx=2, pady=4)

		elif v_cb_nuevo_destino.get() is False and v_cb_destino.get() is True:

			# Destino existente
			# Mostrar los campos de entrada del destino y un botón de revisar	
			b_destino = Button(dobleframe, text="Elegir archivo destino existente: ", command=examinar_destino )
			b_destino.grid(row=16, column=0, sticky="e", padx=2, pady=2)
		

			# botón de confirmar los cambios

			#b_destino = Button(dobleframe, text="Revisar", command=revisar_destino )
			#b_destino.grid(row=posicion_row_grid + 1, column=1, sticky="e", padx=2, pady=4)


		elif v_cb_nuevo_destino.get() is True and v_cb_destino.get() is True: 
			
			# ambos checkboxees marcados
			tk.messagebox.showwarning(title="Destino único", message="Debe elegirse un único destino.")

		else:
			# ambos checkboxees desmarcados
			tk.messagebox.showwarning(title="Elegir destino", message="Debe elegirse un destino.")

		redimensionar_scrollbar()

def nuevo_destino():
	
	a_origen_ok = str(a_origen.replace('\\', '\\\\'))
	wb_origen = load_workbook(filename = a_origen_ok)
	sheet_wb_origen = wb_origen[str(e_sheet_origen.get())]
	#fila_origen=str(e_fila_origen.get())
	fila_origen=str(e_fila_ini_copia_origen.get())

	wb_nuevo_destino = Workbook()
	sheet_wb_nuevo_destino = wb_nuevo_destino.active

	col_ini_origen=str(e_col_ini_origen.get().upper())
	col_fin_origen=str(e_col_fin_origen.get().upper())
	print ("col_ini_origen: " + str(col_ini_origen) + "  col_fin_origen: " + str (col_fin_origen))

	
	for fila in sheet_wb_origen.iter_rows( min_row= int(fila_origen)):
		#print ("fila")
		#print (fila)
		## Copia todas las columnas en nuevo archivo
		#for columna in fila :
		#	#print ("columna")
		#	#print (columna)
		#	coord_destino = (re.search(r"\.(\w+)>$", str(columna)))
		#	celda_destino = str(coord_destino.group(1))
		#	cel_v = columna.value
		#	#print("cel_v")
		#	#print(cel_v)
		#	sheet_wb_nuevo_destino[celda_destino] = cel_v
		# Copia el rango de columnas seleccionado y las pega todas seguidas en el nuevo archivo
		for columna in fila[ord(col_ini_origen) - 65: ord(col_fin_origen) - 65 + 1]:
			celda_destino = columna.coordinate
			cel_v = columna.value
			print ("celda_destino: "+str(celda_destino)+"  cel_v: "+str(cel_v))
			sheet_wb_nuevo_destino[celda_destino] = cel_v

	#wb_nuevo_destino.save('nuevo3.xlsx')
	a_guardar = filedialog.asksaveasfilename(defaultextension=".xlsx")
	wb_nuevo_destino.save(a_guardar)
	tk.messagebox.showinfo(title="¡¡Ya lo tienes!!", message="Los datos se han guardado exitósamente.")

	redimensionar_scrollbar()

def mostrar_destino():

	global wb_destino
	global sheet_wb_destino




	# revisar datos de destino introducidos
	if a_destino == None : 

		tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")

	elif e_sheet_destino == None : 

		tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")

	elif e_fila_destino == None : 

		tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")


	#elif  e_fila_destino == None : 
	#
	#	tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")
	#
	#elif  e_col_ini_destino == None : 
	#
	#	tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")
	#
	#elif  e_col_fin_destino == None : 
	#
	#	tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")

	elif  e_fila_escritura_destino == None : 

		tk.messagebox.showwarning(title="Destino incompleto", message="Deben indicarse todos los parámetros solicitados del destino.")

	else:
		existente_destino()
		l_col_dest = Label(dobleframe,text="Columna donde se copiarán los datos en destino", bg="pale green", wraplength=120)
		l_col_dest.grid(row=26, column=3, sticky="ew", padx=2, pady=2)	

		# elabora las filas del archivo destino existente
		# permite la entry de ordenación de filas
		a_destino_ok = str(a_destino.replace('\\', '\\\\'))
		wb_destino = load_workbook(filename = a_destino_ok)
		# comprobar que exista la pestaña indicada
		pestanas_destino = wb_destino.sheetnames
		if str(e_sheet_destino.get()) in pestanas_destino:
			sheet_wb_destino = wb_destino[str(e_sheet_destino.get())]
			fila_destino=str(e_fila_destino.get())
			col_ini_destino=str(e_col_ini_destino.get())
			col_fin_destino=str(e_col_fin_destino.get())
			#print(str(fila_destino) + " - "+ str(col_ini_destino) + " - "+ str(col_fin_destino))
			celda_ini_destino = str(col_ini_destino + fila_destino)
			celda_fin_destino = str(col_fin_destino + fila_destino)
			rango_celdas_destino = str(celda_ini_destino + ":" + celda_fin_destino)
			#print(rango_celdas_destino)

			# Si se quiere mostrar cabceera de destino
			if e_fila_destino != None and e_col_ini_destino != None and e_col_fin_destino != None :
				tupla_destino = sheet_wb_destino[rango_celdas_destino]
				dic_destino = {}
				#print ("Longitud dic_destino: " + str(len(dic_destino)))
	
				# dicionario columnas de destino: Key=columna_origen y value=columna_destino(donde se desea)
				dic_col_dest = {}
				#print (dic_col_dest)
	
	
				# listar cabecera de destino
		
				posicion_row_grid_destino = 28
		
				#print (tupla_destino)
				for elementos_tupla_destino in tupla_destino :
					#print (elementos_tupla_destino)
					#lista_col_origen.append(elementos_tupla_destino) # No necesario
					#print (lista_col_destino) # No necesario
					# salida: [(<Cell 'Hoja1'.A1>, <Cell 'Hoja1'.B1>, <Cell 'Hoja1'.C1>)]
					for val_col_destino in elementos_tupla_destino:
						# Flecha
						l_flecha = Label(dobleframe,text="-->", bg="pale green", wraplength=30)
						l_flecha.grid(row=posicion_row_grid_destino, column=2, sticky="ew", padx=2, pady=2)
						# separador
						l_separador = Label(dobleframe,text=" | ", bg="pale green", wraplength=30)
						l_separador.grid(row=posicion_row_grid_destino, column=4, sticky="ew", padx=2, pady=2)					
		
						#print (val_col_destino)
						# salida: <Cell 'Hoja1'.A1>
						pos_destino = re.search(r"\.(\w+)>$", str(val_col_destino))
						pos_cel_destino = pos_destino.group(1)
						#pos_col_destino = val_col_destino.column # Las enumera en vez de poner sus letras
						#pos_fil_destino = val_col_destino.row
						#pos_cel_destino = str(pos_col_destino + pos_fil_destino)
						l_pos_col_destino = Label(dobleframe,text=pos_cel_destino, bg="pale green", wraplength=30)
						l_pos_col_destino.grid(row=posicion_row_grid_destino, column=5, sticky="ew", padx=2, pady=2)
						# valores destino
						#print (val_col_destino.value)
						# salida: Col-1
						l_val_col_destino = Label(dobleframe,text=val_col_destino.value, bg="pale green", wraplength=70)
						l_val_col_destino.grid(row=posicion_row_grid_destino, column=6, sticky="w", padx=2, pady=2)
						dic_destino.update({pos_cel_destino:val_col_destino.value})
						#print(dic_destino)
	
						posicion_row_grid_destino = posicion_row_grid_destino + 1
			

				if len(dic_destino) != len(dic_origen):
		
					tk.messagebox.showwarning(title="Diferente tamaño de tablas ", message="Las tablas origen y destino son de diferente tamaño.")

			# Poner los campos enntry para las columnas en destino
			posicion_row_grid_lista_cols_destino = 28
			posicion_row_grid_cols_destino = 28
			# 1-Conocer número de columnas origen
			num_lista_cab_origen = len(lista_cab_origen)
			# 2-Listar nombre de las variables entry (Para luego hacer dicionario de columnas cabecero origen y columnas destino)
			dic_e_col_orig_dest={}
			global lista_e_destino
			lista_e_destino = []
			for n_e_destino in range(num_lista_cab_origen) :
				lista_e_destino.append(Entry(dobleframe))
				lista_e_destino[n_e_destino].grid(row=posicion_row_grid_lista_cols_destino, column=3, sticky="w", padx=2, pady=2)
				posicion_row_grid_lista_cols_destino += 1
				#print(lista_e_destino)


			# Botón para recoger los valores de las columans y confirmar
			b_conf_destino = Button(dobleframe, text="Confirmar en destino existente", command=confirmar_destino )
			b_conf_destino.grid(row=posicion_row_grid + 1, column=1, sticky="e", padx=2, pady=4)	

		else:
			tk.messagebox.showwarning(title="Destino", message="Por favor revisa el nombre de la pestaña de destino.")


		#e_columna_destino = Entry(dobleframe)


	redimensionar_scrollbar()

def existente_destino():
	
	##### Muesta de cabceras de destino

	# Columnas destino
	l_cabecera_destino = Label(dobleframe,text="(Informativo) - Cabecera de DESTINO:", bg="pale green")
	l_cabecera_destino.grid(row=25, column=5, columnspan=2, sticky="ew", padx=4, pady=3)

	l_col_cabe_destino = Label(dobleframe,text="Columnas cabecera destino:", bg="pale green", wraplength=55)
	l_col_cabe_destino.grid(row=26, column=5, sticky="ew", padx=2, pady=2)

	l_val_cabe_destino = Label(dobleframe,text="Título  destino:", bg="pale green", wraplength=70)
	l_val_cabe_destino.grid(row=26, column=6, sticky="w", padx=2, pady=2)

	l_indica_destino = Label(dobleframe,text="Colocar en...", bg="pale green")
	l_indica_destino.grid(row=26, column=2, columnspan=2, sticky="ew", padx=4, pady=3)
	# en la col=2 va "->" y en la col=3 va el entry cpropuesto
	
	##########

	"""
	if cb_nuevo_destino is True:
		wb_origen = load_workbook(filename = a_origen)
		sheet_wb_origen = wb_origen.active
		fila_origen=str(e_fila_origen.get())
		col_ini_origen=int(e_col_ini_origen.get())
		col_fin_origen=int(e_col_fin_origen.get())
		#print(str(fila_origen) + " - "+ str(col_ini_origen) + " - "+ str(col_fin_origen))
		celda_ini_origen = str(col_ini_origen + fila_origen)
		celda_fin_origen = str(col_fin_origen + fila_origen)

	redimensionar_scrollbar()
	"""

	redimensionar_scrollbar()

def confirmar_destino():

	fila_ini_lectura_o = e_fila_ini_copia_origen.get()
	fila_ini_escri_d = e_fila_escritura_destino.get()

	# Obtener celdas de inicio de lectura
	lista_cols_cab_origen = []
	for c_i_l in lista_cab_origen:
		letra = re.search(r"(\w+)\d+", c_i_l)
		c_i_l_o = letra.group(1)
		#celda_ini_lectura_o = str(c_i_l_o + fila_ini_lectura_o )
		lista_cols_cab_origen.append(c_i_l_o)	


	####    Obtener las columnas de destino de los campos entry
	lista_e_cols_destino=[]
	for e_col_d in lista_e_destino :
		lista_e_cols_destino.append(e_col_d.get())
	

	####    Comprobar tamaños
	if len(lista_cols_cab_origen) == len(lista_e_cols_destino):

		# Abrir archivo de origen
		a_origen_ok = str(a_origen.replace('\\', '\\\\'))
		wb_origen = load_workbook(filename = a_origen_ok)
		sheet_wb_origen = wb_origen[str(e_sheet_origen.get())]
		
		# Abrir archivo destino
		a_destino_ok = str(a_destino.replace('\\', '\\\\'))
		wb_destino = load_workbook(filename = a_destino_ok)
		sheet_wb_destino = wb_destino[str(e_sheet_destino.get())]
		
		# Definir la fila de inicio para copiar
		inicio_fila_o = int(fila_ini_lectura_o)
		
		# Definir la fila de destino para escribir
		inicio_fila_d = int(fila_ini_escri_d)

		# Recorrer las filas de origen y copiarlas en las columnas destino (importa el orden de ejecución de filas x columnas)
		for fila_o in range(inicio_fila_o, sheet_wb_origen.max_row + 1):
			for i, columna_origen in enumerate(lista_cols_cab_origen):
				celda_origen = sheet_wb_origen.cell(row=fila_o, column=column_index_from_string(columna_origen))
				columna_destino = lista_e_cols_destino[i]
				sheet_wb_destino.cell(row=inicio_fila_d, column=column_index_from_string(columna_destino), value=celda_origen.value)
			inicio_fila_d += 1 


		wb_destino.save(filename = a_destino_ok)

		tk.messagebox.showinfo(title="¡¡Ya lo tienes!!", message="Los datos se han guardado exitósamente.")



		
	else:
		tk.messagebox.showwarning(title="Diferente tamaño", message="El número de columnas de origen es diferente al de destino.\n Por favor asegúrate que cada columna origen tenga su campo destino")

	redimensionar_scrollbar()






# -------- ESTRUCTURA  --------------

# ---- Raíz ------

root = Tk()
root.title('Unificador de Tablas de Archivos Excel')
root.geometry("700x700")
root.resizable(True, True)
root.iconbitmap("D:/D_lahis/Documents/Python_Scripts/Unificar_tablas_excel/gacela.ico")
#root.config(bg="lightgrey")
barra_menu = Menu(root)
root.config(bg="pale green", bd=2, relief="groove", menu=barra_menu)


# ---- Barra de menú ------

info = Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(menu=info, label="Información")
info.add_command(label="Información", command= lambda : messagebox.showinfo(title="Info", message="1- Elige del archivo origen de Excel la fila y columnas del cabecero de la tabla a copiar.\n\n 2- Posteriormente elige un archivo destino. Éste puede ser un archivo nuevo o bien uno exitente en donde se debe indicar la cabcera de la tabla.\n\n 3- Haz coincidir las columnas de origen con las de destino.\n\n 4- Ejecuta la copia de los valores de la tabla origen a la tabla destino.\n\n 5- Enjoy your time!! "))

sobre = Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(menu=sobre, label="Acerca de...")
sobre.add_command(label="Acerca de..", command= lambda : messagebox.showinfo(title="Acerca de..", message="Aplicación casera para ahorrar tiempo. V.0.2"))

cerrar = Menu(barra_menu, tearoff=False)
barra_menu.add_cascade(menu=cerrar, label="Salir")
cerrar.add_command(label="Salir", command= root.destroy)


# ----- Scrollbar

#scrollbar = tk.Scrollbar(root)
#c = tk.Canvas(root, background="pale green", yscrollcommand=scrollbar.set)
#scrollbar.config(command=c.yview)
#scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

hbar = tk.Scrollbar(root, orient=HORIZONTAL)
vbar = tk.Scrollbar(root, orient=VERTICAL)
c = tk.Canvas(root, background="pale green")
hbar.config(command=c.xview)
hbar.pack(side=tk.BOTTOM,fill=tk.X)
vbar.config(command=c.yview)
vbar.pack(side=tk.RIGHT, fill=tk.Y)
c.config(xscrollcommand=hbar.set, yscrollcommand=vbar.set)
#c.pack(side=LEFT,expand=True,fill=BOTH)
# Para moverse pulsado sobre el panell y no solo sobre las barras.
c.bind_all("<B1-Motion>", funcion_scroll)


# ---- Frame -- Grilla Doble

dobleframe = tk.Frame(c)
dobleframe.pack(fill="both", expand="True")
dobleframe.config(bg="pale green")
dobleframe.config(width=700,height=700)
dobleframe.config(cursor="hand2")


c.pack(side="left", fill="both", expand="True")
c.create_window(4,4,window=dobleframe, anchor="nw")


# TITULO - PRESENTACIÓN

l_presentacion = Label(dobleframe, 
	text="Esta aplicación permite seleccionar columnas de una tabla de un archivo origen\n y copiarlas en un archivo destino.", 
	font=("arial",11, 'bold'), bg="pale green")
l_presentacion.grid(row=1, columnspan=2, sticky="ew", padx=30, pady=10)


# ORIGEN

a_origen=""

# Archivo origen
l_origen = Label(dobleframe,text="1 - Elige archivo origen:", bg="pale green")
l_origen.grid(row=3, sticky="w", columnspan=2, padx=2, pady=2)

#Origen = filedialog.askopenfilename(title="Examinar archivo origen")
b_origen = Button(dobleframe, text="Selecciona archivo origen", command=examinar_origen )
b_origen.grid(row=4, column=0, sticky="e", padx=2, pady=2)

# Datos cabecera de origen
l_origen = Label(dobleframe,text="2 - Elige el rango de fila de la cabecera del archivo origen:", bg="pale green")
l_origen.grid(row=7, sticky="w",columnspan=2, padx=2, pady=2)

l_sheet_origen = Label(dobleframe,text="Indica nombre de pestaña:", bg="pale green")
l_sheet_origen.grid(row=8, column=0, sticky="e", padx=2, pady=2)
e_sheet_origen = Entry(dobleframe)
e_sheet_origen.grid(row=8, column=1, sticky="w", padx=2, pady=2)

l_fila_origen = Label(dobleframe,text="Indica fila de cabecera:", bg="pale green")
l_fila_origen.grid(row=9, column=0, sticky="e", padx=2, pady=2)
e_fila_origen = Entry(dobleframe)
e_fila_origen.grid(row=9, column=1, sticky="w", padx=2, pady=2)

l_col_ini_origen = Label(dobleframe,text="Indica columna de inicio:", bg="pale green")
l_col_ini_origen.grid(row=10, column=0, sticky="e", padx=2, pady=2)
e_col_ini_origen = Entry(dobleframe)
e_col_ini_origen.grid(row=10, column=1, sticky="w", padx=2, pady=2)

l_col_fin_origen = Label(dobleframe,text="Indica columna de fin:", bg="pale green")
l_col_fin_origen.grid(row=11, column=0, sticky="e", padx=2, pady=2)
e_col_fin_origen = Entry(dobleframe)
e_col_fin_origen.grid(row=11, column=1, sticky="w", padx=2, pady=2)

l_fila_ini_copia_origen = Label(dobleframe,text="Indica fila En la que quieres empezar a copiar los datos - (INCLUIDA):", bg="pale green")
l_fila_ini_copia_origen.grid(row=12, column=0, sticky="e", padx=2, pady=2)
e_fila_ini_copia_origen = Entry(dobleframe)
e_fila_ini_copia_origen.grid(row=12, column=1, sticky="w", padx=2, pady=2)

# DESTINO

# Archivo destino
l_destino = Label(dobleframe,text="3 - Elige archivo destino:", bg="pale green")
l_destino.grid(row=13, column=0, sticky="w", padx=2, pady=2)

l_nuevo_destino = Label(dobleframe,text="¿Quieres crear un nuevo archivo?", bg="pale green")
l_nuevo_destino.grid(row=14, column=0, sticky="e", padx=2, pady=2)
v_cb_nuevo_destino = BooleanVar()
cb_nuevo_destino = Checkbutton(dobleframe, variable=v_cb_nuevo_destino)
cb_nuevo_destino.grid(row=14, column=1, sticky="w", padx=2, pady=2)

l_destino = Label(dobleframe,text="   ¿Quieres utilizar un archivoexistente?", bg="pale green")
l_destino.grid(row=15, column=0, sticky="e", padx=2, pady=2)
v_cb_destino = BooleanVar()
cb_destino = Checkbutton(dobleframe, variable=v_cb_destino)
cb_destino.grid(row=15, column=1, sticky="w", padx=2, pady=2)


# REVISAR COLUMNAS

b_destino = Button(dobleframe, text="Revisar columnas", command=revisar_columnas )
b_destino.grid(row=25, column=0, sticky="w", padx=30, pady=10)

# Columnas origen
l_cabecera_origen = Label(dobleframe,text="Cabecera de ORIGEN:", bg="pale green")
l_cabecera_origen.grid(row=26, column=0, columnspan=2, sticky="ew", padx=4, pady=3)

l_col_cabe_origen = Label(dobleframe,text="Columnas origen:", bg="pale green", wraplength=55)
l_col_cabe_origen.grid(row=27, column=0, sticky="ew", padx=2, pady=2)

l_val_cabe_origen = Label(dobleframe,text="Título  origen:", bg="pale green", wraplength=70)
l_val_cabe_origen.grid(row=27, column=1, sticky="w", padx=2, pady=2)




redimensionar_scrollbar()
#root.update()
#c.config(scrollregion=c.bbox("all"))

root.mainloop()
